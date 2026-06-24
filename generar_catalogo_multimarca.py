# -*- coding: utf-8 -*-
"""
Genera catalogo.json multi-marca a partir de Bd.xlsx.
Preserva IDs existentes de NutriVita (60 productos, categorías 1-3, subcategorías 1-25).

Para cada producto almacena `imagen` con la ruta exacta relativa al container de blob,
detectada escaneando Imagenes/Catalogo_Productos/ (preserva extensión y estructura real).

Reglas especiales NutriVita (mismo que sync_imagenes_blob.py):
  - MELATONINA.{ext} → NV1670.{ext}
  - Medicamentos/Articular/NV1691* → Fitoterapéuticos/Articular/NV1691*
"""
import sys, os, json
import openpyxl
sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
EXCEL_SRC  = os.path.join(BASE_DIR, 'Bd.xlsx')
JSON_OUT   = os.path.join(BASE_DIR, 'catalogo.json')
IMG_ROOT   = os.path.join(BASE_DIR, 'Imagenes', 'Catalogo_Productos')

MARCAS = [
    {'id': 1, 'nombre': 'NutriVita',  'blob_carpeta': 'NutriVita'},
    {'id': 2, 'nombre': 'Nutra',      'blob_carpeta': 'Nutra'},
    {'id': 3, 'nombre': 'GreenLab',   'blob_carpeta': 'GreenLab'},
    {'id': 4, 'nombre': 'Boiron',     'blob_carpeta': 'Boiron'},
    {'id': 5, 'nombre': 'Atopeel',    'blob_carpeta': 'Atopel'},
]

HOJA_A_MARCA_ID = {
    'NutriVita': 1, 'Nutra': 2, 'GreenLab': 3, 'Boiron': 4, 'Atopeel': 5,
}

HOJA_COLS = {
    'NutriVita': {'cat':0,'sub':1,'nom':2,'cod':3,'inv':4,'obs':5,'pre':6},
    'Nutra':     {'cat':0,'sub':1,'nom':2,'cod':3,'inv':4,'obs':5,'pre':6},
    'GreenLab':  {'cat':0,'sub':1,'nom':2,'cod':3,'inv':None,'obs':5,'pre':None},
    'Boiron':    {'cat':0,'sub':1,'nom':2,'cod':3,'inv':4,'obs':5,'pre':None},
    'Atopeel':   {'cat':0,'sub':1,'nom':2,'cod':3,'inv':None,'obs':5,'pre':None},
}

# Renombres que aplica sync_imagenes_blob.py (local relativo → blob relativo)
RENAME_RULES = {
    'NutriVita/Medicamentos/Sistema Nervioso/MELATONINA.png':
        'NutriVita/Medicamentos/Sistema Nervioso/NV1670.png',
    'NutriVita/Medicamentos/Sistema Nervioso/MELATONINA_01.png':
        'NutriVita/Medicamentos/Sistema Nervioso/NV1670_01.png',
    'NutriVita/Medicamentos/Articular/NV1691.png':
        'NutriVita/Fitoterapéuticos/Articular/NV1691.png',
    'NutriVita/Medicamentos/Articular/NV1691_01.png':
        'NutriVita/Fitoterapéuticos/Articular/NV1691_01.png',
}


def construir_indice_imagenes():
    """
    Escanea Imagenes/Catalogo_Productos/ y construye:
      blob_path_by_code[(blob_carpeta, codigo_sin_ext)] = blob_relative_path

    Solo indexa imágenes PRINCIPALES (sin sufijo _01, _1, etc.).
    """
    idx = {}
    if not os.path.isdir(IMG_ROOT):
        print(f'  ⚠ Directorio de imágenes no encontrado: {IMG_ROOT}')
        return idx

    for root, _dirs, files in os.walk(IMG_ROOT):
        for fname in files:
            full   = os.path.join(root, fname)
            rel    = os.path.relpath(full, IMG_ROOT).replace('\\', '/')
            blob   = RENAME_RULES.get(rel, rel)  # aplica renombres

            # Separar código y extensión
            name_no_ext, ext = os.path.splitext(os.path.basename(blob))

            # Ignorar variantes (_01, _1, _02...)
            import re
            if re.search(r'_\d+$', name_no_ext):
                continue

            # Primera parte del blob path = blob_carpeta (marca)
            parts = blob.split('/', 1)
            blob_carpeta = parts[0]

            # Guardar en índice: (blob_carpeta, codigo) → blob_path
            key = (blob_carpeta, name_no_ext)
            if key not in idx:
                idx[key] = blob
    return idx


def main():
    # ── Índice de imágenes locales → blob path ────────────────────────────────
    img_idx = construir_indice_imagenes()
    print(f'  Imágenes indexadas: {len(img_idx)}')

    def obtener_imagen(blob_carpeta, codigo):
        key = (blob_carpeta, str(codigo))
        blob = img_idx.get(key)
        if blob:
            return blob
        # Fallback: intentar con código como entero (p.ej. "04" → "4")
        try:
            key2 = (blob_carpeta, str(int(codigo)))
            blob = img_idx.get(key2)
            if blob:
                return blob
        except ValueError:
            pass
        return None  # No encontrada

    # ── Cargar JSON actual — solo productos NutriVita para preservar sus IDs ──
    with open(JSON_OUT, encoding='utf-8') as f:
        catalogo_actual = json.load(f)

    # Tomar SOLO productos NutriVita (marca_id==1 o sin campo marca_id = JSON original)
    nutrivita_prods = [
        p for p in catalogo_actual['productos']
        if p.get('marca_id', 1) == 1 or p.get('marca') == 'NutriVita'
        or ('marca_id' not in p and 'marca' not in p)
    ]

    # Para IDs iniciales: basarse en los IDs de las categorías NutriVita existentes
    # (pueden estar mezcladas con otras marcas si el JSON ya fue actualizado)
    nutrivita_cats = [c for c in catalogo_actual['categorias']
                      if c.get('marca_id', 1) == 1 or c.get('marca') == 'NutriVita'
                      or 'marca_id' not in c]
    nutrivita_subs = [s for c in nutrivita_cats for s in c.get('subcategorias', [])]

    existing_cat_by_name = {c['nombre']: c['id'] for c in nutrivita_cats}
    existing_sub_by_name = {s['nombre']: s['id'] for s in nutrivita_subs}

    max_nutrivita_prod_id = max((p['id'] for p in nutrivita_prods), default=0)
    max_nutrivita_cat_id  = max((c['id'] for c in nutrivita_cats),  default=0)
    max_nutrivita_sub_id  = max((s['id'] for s in nutrivita_subs),  default=0)

    next_id = {
        'prod': max_nutrivita_prod_id + 1,
        'cat':  max_nutrivita_cat_id  + 1,
        'sub':  max_nutrivita_sub_id  + 1,
    }

    # ── Inicializar con NutriVita existentes ──────────────────────────────────
    cats_index = {}
    for c in nutrivita_cats:
        key = (1, c['nombre'])
        cats_index[key] = {
            'id': c['id'], 'nombre': c['nombre'], 'marca_id': 1,
            'marca': 'NutriVita', 'subcats': {},
        }
        for s in c.get('subcategorias', []):
            cats_index[key]['subcats'][(1, s['nombre'])] = {
                'id': s['id'], 'nombre': s['nombre']
            }

    # Copiar solo los 60 productos NutriVita con imagen actualizada
    productos = []
    existing_nutrivita_codes = set()
    for p in nutrivita_prods:
        p2 = dict(p)
        p2['marca_id']     = 1
        p2['marca']        = 'NutriVita'
        p2['blob_carpeta'] = 'NutriVita'
        img = obtener_imagen('NutriVita', p['codigo'])
        p2['imagen'] = img if img else None
        productos.append(p2)
        existing_nutrivita_codes.add(p['codigo'])

    # ── Procesar cada hoja del Excel ─────────────────────────────────────────
    wb = openpyxl.load_workbook(EXCEL_SRC, data_only=True)

    for hoja_nombre, marca_id in HOJA_A_MARCA_ID.items():
        if hoja_nombre not in wb.sheetnames:
            print(f'  ⚠ Hoja "{hoja_nombre}" no encontrada.')
            continue

        ws         = wb[hoja_nombre]
        cols       = HOJA_COLS[hoja_nombre]
        marca_info = MARCAS[marca_id - 1]
        added = skipped = 0

        for row in range(2, ws.max_row + 1):
            def cv(col_idx, r=row):
                if col_idx is None: return None
                v = ws.cell(r, col_idx + 1).value
                return str(v).strip() if v is not None else None

            cat_nombre = cv(cols['cat'])
            sub_nombre = cv(cols['sub'])
            nom        = cv(cols['nom'])
            if not nom or not cat_nombre or not sub_nombre:
                continue

            cod_raw = ws.cell(row, cols['cod'] + 1).value
            if cod_raw is None:
                print(f'  ⚠ [{hoja_nombre}] fila {row}: sin código para "{nom}".')
                continue
            codigo = str(int(cod_raw)) if isinstance(cod_raw, (int, float)) else str(cod_raw).strip()

            inv = cv(cols['inv']) if cols['inv'] is not None else None
            obs = cv(cols['obs'])
            precio = 0.0
            if cols.get('pre') is not None:
                pre_raw = ws.cell(row, cols['pre'] + 1).value
                if pre_raw is not None:
                    try: precio = float(pre_raw)
                    except (ValueError, TypeError): precio = 0.0

            if marca_id == 1 and codigo in existing_nutrivita_codes:
                skipped += 1
                continue

            # Categoría
            cat_key = (marca_id, cat_nombre)
            if cat_key not in cats_index:
                if marca_id == 1 and cat_nombre in existing_cat_by_name:
                    cat_id = existing_cat_by_name[cat_nombre]
                else:
                    cat_id = next_id['cat']; next_id['cat'] += 1
                cats_index[cat_key] = {
                    'id': cat_id, 'nombre': cat_nombre,
                    'marca_id': marca_id, 'marca': marca_info['nombre'], 'subcats': {},
                }

            cat_entry = cats_index[cat_key]

            # Subcategoría
            sub_key = (marca_id, sub_nombre)
            if sub_key not in cat_entry['subcats']:
                if marca_id == 1 and sub_nombre in existing_sub_by_name:
                    sub_id = existing_sub_by_name[sub_nombre]
                else:
                    sub_id = next_id['sub']; next_id['sub'] += 1
                cat_entry['subcats'][sub_key] = {'id': sub_id, 'nombre': sub_nombre}

            sub_entry = cat_entry['subcats'][sub_key]

            # Imagen
            img = obtener_imagen(marca_info['blob_carpeta'], codigo)

            p_id = next_id['prod']; next_id['prod'] += 1
            productos.append({
                'id':             p_id,
                'codigo':         codigo,
                'nombre':         nom,
                'categoria':      cat_nombre,
                'categoria_id':   cat_entry['id'],
                'subcategoria':   sub_nombre,
                'subcategoria_id': sub_entry['id'],
                'registro_invima': inv,
                'precio':         precio,
                'observaciones':  obs if obs else None,
                'activo':         True,
                'marca':          marca_info['nombre'],
                'marca_id':       marca_id,
                'blob_carpeta':   marca_info['blob_carpeta'],
                'imagen':         img,
            })
            added += 1

        print(f'  {hoja_nombre}: {added} nuevos, {skipped} ya existían.')

    # ── Lista de categorías ───────────────────────────────────────────────────
    cats_list = []
    for mid in [1, 2, 3, 4, 5]:
        for _, e in sorted(
            ((k, v) for k, v in cats_index.items() if v['marca_id'] == mid),
            key=lambda kv: kv[1]['id']
        ):
            cats_list.append({
                'id':          e['id'],
                'nombre':      e['nombre'],
                'marca_id':    e['marca_id'],
                'marca':       e['marca'],
                'subcategorias': sorted(e['subcats'].values(), key=lambda s: s['id']),
            })

    # ── Escribir JSON ─────────────────────────────────────────────────────────
    catalogo_nuevo = {
        'marcas':     MARCAS,
        'categorias': cats_list,
        'productos':  sorted(productos, key=lambda p: p['id']),
    }

    with open(JSON_OUT, 'w', encoding='utf-8') as f:
        json.dump(catalogo_nuevo, f, ensure_ascii=False, indent=2)

    sin_imagen = sum(1 for p in productos if not p.get('imagen'))
    print(f'\n✅ catalogo.json: {len(MARCAS)} marcas | {len(cats_list)} categorías | {len(productos)} productos')
    print(f'   Sin imagen encontrada: {sin_imagen}')
    print(f'   Guardado en: {JSON_OUT}')


if __name__ == '__main__':
    main()
