# -*- coding: utf-8 -*-
"""
Actualiza el archivo de auditoria NutriHeal 360:
1. Reemplaza los prompts de imagen (columna 9) con la plantilla optimizada
   para ElevenLabs Image que preserva el producto real.
2. Agrega dos columnas nuevas con hipervínculos a las imágenes de referencia
   que el operador debe subir al generador de imágenes:
      - Col 13: Imagen Producto (Ref. 1) → URL Azure Blob
      - Col 14: Logo NutriHeal    (Ref. 2) → archivo local
3. Regenera las hojas Publer Import con los alt-texts actualizados.
"""
import json
import os
import re
import sys
import urllib.parse
import csv

sys.stdout.reconfigure(encoding='utf-8')

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_XLSX = os.path.join(BASE_DIR, 'NutriHeal360_DrMestizo_Auditoria_Contenido_3Meses_2026.xlsx')
CATALOGO_JSON = os.path.join(BASE_DIR, 'catalogo.json')

BLOB_BASE = 'https://nutriheal.blob.core.windows.net/productos/'
LOGO_PATH_LOCAL = os.path.join(BASE_DIR, 'Imagenes', 'logo', 'logo.png')
LOGO_URL = 'file:///' + LOGO_PATH_LOCAL.replace('\\', '/').replace('G:/', 'G:/')
WHATSAPP_LINK = 'https://wa.me/573147088080'

# ── Carga catálogo ──────────────────────────────────────────────────────────
with open(CATALOGO_JSON, encoding='utf-8') as f:
    catalog_data = json.load(f)

CATALOG_MAP = {
    p['nombre'].upper().strip(): p
    for p in catalog_data['productos']
}

def blob_url_producto(nombre):
    """Devuelve la URL pública de la imagen principal en Azure Blob Storage."""
    key = nombre.upper().strip()
    if key not in CATALOG_MAP:
        return None
    p = CATALOG_MAP[key]
    # Usar campo imagen exacto si existe (multi-marca, extensión correcta)
    if p.get('imagen'):
        return BLOB_BASE + '/'.join(urllib.parse.quote(seg) for seg in p['imagen'].split('/'))
    # Fallback legacy (solo NutriVita, estructura jerárquica)
    cat    = urllib.parse.quote(p['categoria'])
    sub    = urllib.parse.quote(p['subcategoria'])
    codigo = p['codigo']
    return f'{BLOB_BASE}{cat}/{sub}/{codigo}.png'

# ── Plantillas de prompts ────────────────────────────────────────────────────
SUPLEMENTACION_TEMPLATE = (
    # ── SUBJECT (reference-locked) ──────────────────────────────────────────
    'Photorealistic luxury product photograph. '
    'SUBJECT: the supplement bottle labeled "{product}" as shown in Reference Image 1 — '
    'reproduce its label artwork, color palette, typography, capsule/bottle shape, cap design, '
    'and all branding elements with 100% pixel fidelity. '
    'Do NOT alter, redesign, recolor, or reimagine the label in any way. '
    # ── SURFACE & BACKGROUND ────────────────────────────────────────────────
    'SURFACE: polished deep-black Marquina marble with fine white veining, '
    'photographed straight-on at a slight 10° downward camera tilt. '
    'BACKGROUND: smooth dark charcoal-to-black gradient bokeh, '
    'with a diffuse emerald green (#1B7A4A) ambient halo emanating from behind the bottle, '
    'intensity 30% — creates depth without competing with the label. '
    # ── LIGHTING ────────────────────────────────────────────────────────────
    'LIGHTING: three-point studio setup — '
    '(1) large softbox camera-left at 45°, 70% intensity, warm white 5500K; '
    '(2) gold reflector (#C9A84C) rim light from camera-right, 25% intensity, '
    'creates a fine warm edge highlight along the bottle contour; '
    '(3) subtle back-light from below the marble, emerald green gel, 15% intensity. '
    'No harsh shadows. Bottle casts a soft elongated reflection on the marble. '
    # ── FOREGROUND STYLING ───────────────────────────────────────────────────
    'FOREGROUND BOTANICALS: naturally arranged at the bottle base — '
    'one halved orange showing vivid citrus cross-section, '
    'one small knob of fresh ginger root with skin texture visible, '
    'two sprigs of rosemary or thyme with fine needle detail, '
    'a small scatter of raw seeds or dried berries. '
    'All botanicals are sharp and vibrant; none obscure the product label. '
    # ── CAMERA & COMPOSITION ────────────────────────────────────────────────
    'CAMERA: 85mm portrait lens equivalent, f/2.2 aperture. '
    'Bottle occupies 55% of frame height, centered horizontally, '
    'lower third of frame has margin for text overlay if needed. '
    'Shallow depth of field: label tack-sharp, botanicals and background softly blurred. '
    # ── LOGO PLACEMENT ──────────────────────────────────────────────────────
    'LOGO: insert the NutriHeal 360 logo from Reference Image 2 in the bottom-right corner. '
    'Reproduce every element exactly — green leaf icon, "NutriHeal 360" wordmark, '
    'original color values, original proportions. '
    'Logo size: 11–13% of total frame width. '
    'Position: 2% margin from right edge, 2% margin from bottom edge. '
    'Logo must NOT overlap the bottle or any botanical element. '
    'No glow, shadow, or color shift applied to the logo. '
    # ── MOOD & OUTPUT ───────────────────────────────────────────────────────
    'MOOD: pharmaceutical precision meets artisan wellness. '
    'High-end Colombian nutraceutical brand. Clean, confident, trustworthy. '
    'Output: 1:1 square, minimum 2048×2048 px, sRGB color space, no watermarks.'
)

LOGO_TAG = (
    ' LOGO: insert the NutriHeal 360 logo from the reference image in the bottom-right corner — '
    'reproduce every element exactly (green leaf icon, "NutriHeal 360" wordmark, '
    'original colors, original proportions). '
    'Size: 11–13% of frame width. Position: 2% margin from right and bottom edges. '
    'No glow, shadow, or recoloring applied to the logo. '
    'Output: 1:1 square, minimum 2048×2048 px, sRGB, no watermarks.'
)

OLD_TO_NEW_PROMPT = {
    # ── Medicina Alternativa ────────────────────────────────────────────────

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, homeopathic glass vials and pellets on botanical background, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        (
            'Photorealistic still-life photograph for a luxury Colombian wellness brand. '
            'SUBJECT: three to four small amber glass homeopathic vials (10ml dropper bottles) '
            'and a scatter of tiny white pellets arranged on a dark walnut surface. '
            'One vial is uncapped and slightly tilted; fine pellets spill from its opening. '
            'BOTANICALS: dried chamomile flowers, a sprig of fresh arnica, and two small dried '
            'calendula petals placed naturally among the vials — no artificial arrangement. '
            'BACKGROUND: blurred warm consultation room interior, bookshelves with medical texts '
            'barely visible, very soft depth of field. '
            'LIGHTING: single large softbox camera-left 5500K; warm gold (#C9A84C) reflector '
            'camera-right adds a thin rim; emerald green (#1B7A4A) gel backlight 15% intensity '
            'creates a subtle halo behind the vials. '
            'MOOD: scientific trust meets artisan medicine — calm, precise, premium. '
            'No text overlays, no digital effects, no stock-photo feel. '
            'Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, iris of human eye close-up revealing health patterns, iridology diagnostic chart overlay, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        (
            'Ultra-close-up macro photograph of a human iris in extreme detail. '
            'EYE: dark brown iris typical of a Colombian adult; fine radiating trabecula fibers '
            'clearly visible; pupil sharp and dark. '
            'OVERLAY: a semi-transparent iridology chart (emerald green #1B7A4A lines, '
            'opacity 35%) mapped precisely onto the iris surface — organ zones labeled in '
            'fine gold (#C9A84C) sans-serif text, letters no larger than the trabecula fibers. '
            'LIGHTING: ring flash centered on the iris, 5500K, even and shadow-free; '
            'a faint emerald green reflection in the pupil adds depth. '
            'BACKGROUND: blurred circular bokeh in deep teal-green, evoking both technology '
            'and nature. '
            'DEPTH OF FIELD: iris completely sharp edge-to-edge; eyelashes softly blurred. '
            'MOOD: diagnostic precision, natural intelligence, holistic science. '
            'No eyelid makeup, no artificial color grading that changes the iris hue. '
            'Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, vintage botanical medical illustration with modern gold frame overlay, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        (
            'Editorial-style photograph of a hand-painted vintage botanical medical illustration '
            'mounted inside a polished gold (#C9A84C) frame. '
            'ILLUSTRATION: 19th-century style, depicting a medicinal plant (valerian, ginkgo, '
            'or echinacea) with roots, stem, leaves, and flower labeled in old-Latin calligraphy. '
            'Sepia-toned paper showing age, hand-watercolored in muted botanical greens and '
            'ochre tones. '
            'FRAME: ornate brushed gold metal, double-moulding, subtle patina — photographed '
            'against a dark background so the gold glows. '
            'LIGHTING: raking light from camera-right at 30° to reveal the paper texture; '
            'soft emerald green (#1B7A4A) ambient light from behind the frame creates depth. '
            'BACKGROUND: blurred dark emerald velvet or aged wood — no distracting elements. '
            'MOOD: heritage wisdom meeting modern luxury healthcare. '
            'Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, human silhouette with glowing energy meridian lines, holistic healing visualization, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        (
            'Conceptual wellness visualization: a semi-transparent human body silhouette '
            '(slim adult figure, gender-neutral, viewed from front) rendered in translucent '
            'deep teal, floating centered in a dark charcoal environment. '
            'MERIDIAN LINES: 8–10 flowing energy pathways trace the traditional TCM meridian '
            'routes — rendered as luminous emerald green (#1B7A4A) lines with soft particle '
            'glow, 2–3 px wide, with animated-style directional light flow suggested by '
            'varying brightness along each line. '
            'KEY NODES: 7 circular gold (#C9A84C) energy points (chakra/meridian intersections) '
            'pulse with a soft outer glow — each 12px diameter sphere. '
            'BACKGROUND: deep space-like darkness with subtle emerald green particles drifting '
            'upward, creating a sense of energy and ascent. '
            'STYLE: 3D CGI realism, photorealistic skin-like translucency, cinematic grade. '
            'No text, no labels, no anatomical annotations. '
            'MOOD: healing energy, natural intelligence, holistic medicine. '
            'Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,

    # ── Prevención Médica Diaria ────────────────────────────────────────────

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, fresh organic vegetables and medicinal plants arranged on dark slate, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        (
            'Top-down flat-lay photograph on a dark charcoal slate surface. '
            'PRODUCE: a curated selection of 12–15 fresh whole vegetables and medicinal herbs — '
            'red tomatoes, broccoli florets, baby spinach, purple kale, half an avocado, '
            'raw garlic cloves, fresh turmeric root (cut open revealing orange interior), '
            'a halved lemon, fresh ginger, parsley sprigs, and one whole pomegranate. '
            'All items are at peak freshness, vibrant color, no wilting or blemishes. '
            'ARRANGEMENT: organic, editorial food-styling — not a perfect grid, but a natural '
            'abundant spread with deliberate negative space in the lower-right quadrant '
            'for the NutriHeal 360 logo. '
            'LIGHTING: soft overhead diffused natural light (10am window equivalent), '
            '5600K; emerald green (#1B7A4A) color cast in the shadows. '
            'Small gold (#C9A84C) accent props: two tiny ceramic medicine vials placed '
            'among the vegetables. '
            'MOOD: nutritional abundance, Colombian wellness culture, hope and vitality. '
            'Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, human body anatomical illustration with glowing healthy organs, emerald green highlights, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        (
            'Medical-grade 3D CGI render of a human torso cross-section, gender-neutral, '
            'showing healthy internal organs in photorealistic anatomical detail. '
            'ORGANS: heart (slightly enlarged, vivid red with visible coronary vessels), '
            'lungs (full and pink), liver (healthy amber), kidneys (both visible), '
            'stomach and intestines (simplified but accurate). '
            'GLOW EFFECT: each organ emits a subtle emerald green (#1B7A4A) inner-light pulse — '
            'green overlay at 20% opacity — symbolizing peak health. '
            'LABELS: fine gold (#C9A84C) connecting lines point to each organ with a clean '
            'white sans-serif label (12pt equivalent). '
            'BACKGROUND: dark neutral gray-charcoal gradient, no medical environment visible. '
            'SKIN: translucent torso boundary visible as a thin frosted-glass layer. '
            'STYLE: Zygote Body / Visible Human quality; cinematic lighting with soft rim. '
            'MOOD: health education, scientific confidence, preventive medicine. '
            'Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, healthy Colombian family at golden hour outdoor setting, vibrant wellness energy, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        (
            'Golden-hour lifestyle photograph of a healthy Colombian family of three — '
            'parents (30s–40s) and one child (8–12 years old) — outdoors in a lush green park. '
            'MOMENT: candid mid-activity — the child runs slightly ahead while parents walk '
            'closely together, all laughing naturally. No forced poses. '
            'DEMOGRAPHICS: distinctly Colombian features — warm mestizo skin tones, '
            'dark hair, everyday athletic wear in earth tones. '
            'LIGHT: 6:00 pm golden hour, sun at 15° above horizon camera-right; '
            'long warm shadows, skin glows with golden (#C9A84C) backlight, '
            'green grass reflects a faint emerald (#1B7A4A) fill light on shadowed side. '
            'BACKGROUND: Cali urban park — tall ceiba trees in background, '
            'distant Farallones mountains barely visible in the haze. '
            'LENS: 85mm f/1.8, subject sharp, background beautifully bokeh. '
            'MOOD: achievable wellness, family vitality, Colombian pride, hope. '
            'No stock-photo feel. Natural, editorial, authentic. '
            'Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, DNA helix structure with protective shield overlay, prevention concept, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        (
            'Scientific CGI visualization: a photorealistic DNA double helix floating in '
            'a dark void, occupying 60% of the frame height, slightly tilted 15° on its '
            'vertical axis for dynamism. '
            'HELIX DETAIL: full base-pair rungs visible; outer backbone strands rendered as '
            'deep emerald green (#1B7A4A) metallic ribbons with specular highlight; '
            'base pairs in alternating muted gold (#C9A84C) and white. '
            'SHIELD: a semi-transparent hexagonal force-shield surrounding the helix — '
            'thin emerald green lines forming the honeycomb lattice, opacity 40%, '
            'with a soft electric pulse effect at the perimeter. '
            'PARTICLE FIELD: fine glowing particles (emerald and gold) drift upward '
            'around the helix — suggests energy, protection, and dynamism. '
            'LIGHTING: directional cinematic key from camera-left; '
            'cool blue-teal ambient fill from below. '
            'MOOD: genetic-level protection, cutting-edge preventive science, confidence. '
            'No text overlays. Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,

    # ── Casos de Éxito / Trayectoria ───────────────────────────────────────

    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, Colombian patient and doctor moment of success in medical consultation room, warm smile, trust and gratitude emotion, San Fernando Cali medical office aesthetic, certificate wall background, emerald green plants, gold framed diplomas, before-after transformation concept':
        (
            'Warm editorial photograph inside a private medical consultation room in '
            'San Fernando, Cali, Colombia. '
            'SUBJECT: a Colombian male doctor (50s, silver-streaked dark hair, white coat) '
            'facing a smiling female patient (40s, warm mestizo complexion, business-casual). '
            'MOMENT: the doctor extends his right hand for a congratulatory handshake; '
            'both are mid-laugh — natural, unscripted, genuine trust. '
            'ROOM DETAILS: behind the doctor — a wall of gold-framed academic diplomas '
            'and INVIMA certificates (Latin text partially visible); '
            'a healthy monstera plant to the left in a terracotta pot; '
            'a dark-wood desk with a closed laptop and a branded NutriHeal notepad. '
            'LIGHTING: warm overhead tungsten 3200K; a floor lamp behind the patient '
            'provides soft emerald (#1B7A4A) upward fill that reflects off the white walls. '
            'Gold diploma frames catch the warm light and add specular glints. '
            'LENS: 50mm f/2.0, subjects sharp, room detail visible but softly blurred. '
            'MOOD: transformation achieved, trust earned, medical authority combined with '
            'human warmth. Colombian clinical excellence. '
            'Aspect ratio 1:1, minimum 2048×2048 px.'
        ) + LOGO_TAG,
}

# Regex para detectar prompts de Suplementación (formato viejo)
PRODUCTO_RE_OLD = re.compile(r'of "([^"]+)" on a dark marble')
# Regex para detectar prompts de Suplementación (formato nuevo)
PRODUCTO_RE_NEW = re.compile(r'Keep the supplement bottle labeled (.+?) exactly as shown')

# Textos de logo ANTERIORES que deben reemplazarse (formato intermedio)
_OLD_LOGO_SIMPLE = (
    ' Place the NutriHeal 360 logo from the reference image as a small brand watermark '
    'in the bottom-right corner of the composition.'
)
_OLD_LOGO_SUPL = (
    'Place the NutriHeal 360 logo from the second reference image as a small separate brand '
    'watermark in the bottom-right corner of the composition, clearly outside the product and '
    'not overlapping or touching the bottle itself.'
)

def nuevo_prompt(prompt_viejo):
    """Devuelve el prompt actualizado y el nombre del producto si aplica."""
    if not prompt_viejo:
        return prompt_viejo, None

    # Suplementación — formato ORIGINAL
    m = PRODUCTO_RE_OLD.search(prompt_viejo)
    if m:
        nombre = m.group(1)
        return SUPLEMENTACION_TEMPLATE.format(product=nombre), nombre

    # No-Suplementación — formato ORIGINAL (viejo prompt de medical marketing)
    nuevo = OLD_TO_NEW_PROMPT.get(prompt_viejo)
    if nuevo:
        return nuevo, None

    # Suplementación — formato NUEVO pero con logo viejo (re-ejecución)
    m2 = PRODUCTO_RE_NEW.search(prompt_viejo)
    if m2:
        nombre = m2.group(1)
        return SUPLEMENTACION_TEMPLATE.format(product=nombre), nombre

    # No-Suplementación — formato intermedio (wellness photograph + logo viejo)
    # Solo reemplazar la instrucción del logo por la nueva versión estricta
    if _OLD_LOGO_SIMPLE in prompt_viejo:
        return prompt_viejo.replace(_OLD_LOGO_SIMPLE, LOGO_TAG), None

    return prompt_viejo, None

# ── Generación dinámica de instrucción HeyGen ────────────────────────────────

_SKIP_HEYGEN = ('📲', '📍', '🔴', '💬', '👇', '👍', '🔔', '⏱', '#', 'wa.me',
                'WhatsApp', 'www.', 'doctormestizo', 'Cra ', '+57')

def _limpiar_linea(line):
    return re.sub(r'^[^\w¿"]+', '', line).strip()

def extraer_titulo_heygen(copy, pilar):
    """Extrae el tema/título del post para la instrucción HeyGen."""
    if not copy:
        return 'Salud y bienestar con NutriHeal 360'
    lines = [l.strip() for l in copy.split('\n') if l.strip()]

    # YouTube/Facebook: título entre comillas en las primeras 3 líneas
    for line in lines[:3]:
        m = re.search(r'"([^"]{15,})"', line)
        if m:
            return m.group(1).strip()

    # Prevención: "PREVENCIÓN HOY: Tema"
    for line in lines[:2]:
        m = re.search(r'PREVENCIÓN HOY\s*:\s*(.+)', line, re.IGNORECASE)
        if m:
            return m.group(1).strip()

    # Casos de Éxito: línea tras "HISTORIA DE TRANSFORMACIÓN"
    for i, line in enumerate(lines[:3]):
        if 'HISTORIA DE' in line.upper() or 'TRAYECTORIA' in line.upper():
            for j in range(i+1, min(i+3, len(lines))):
                cand = _limpiar_linea(lines[j])
                if len(cand) > 10 and not any(s in lines[j] for s in _SKIP_HEYGEN):
                    return cand
            break

    # Suplementación: "Hoy hablo de PRODUCTO — nuestra línea SUBCATEGORIA"
    if 'Suplementación' in pilar:
        for line in lines:
            m = re.search(r'(?:hablo de|habla de)\s+([A-ZÁÉÍÓÚÑÜ][A-Z0-9\- ]{2,}[A-Z0-9])', line)
            if m:
                prod = m.group(1).strip()
                sm = re.search(r'línea\s+(\w[\w\s]+?)\.', line)
                subcat = f' — línea {sm.group(1).strip()}' if sm else ''
                return f'Beneficios de {prod}{subcat}'

    # Fallback: primera línea limpia
    return _limpiar_linea(lines[0])[:120]


def extraer_puntos_guion(copy, n=3):
    """Extrae hasta n oraciones sustantivas del copy para el guión."""
    if not copy:
        return []
    lines = [l.strip() for l in copy.split('\n') if l.strip()]
    puntos = []
    for line in lines[1:]:
        if any(s in line for s in _SKIP_HEYGEN):
            continue
        # Saltar líneas que son solo un título entre comillas (ya va en TEMA)
        if re.match(r'^"[^"]{10,}"\.?$', line.strip()):
            continue
        clean = re.sub(r'^[🌿🔬🛡️🎙️✅✨💚🧬🌱💊🧪•◦▸→\-\*✓]+\s*', '', line).strip()
        clean = re.sub(r'[^\w\sáéíóúÁÉÍÓÚñÑüÜ:.,¿?!""()\-]', '', clean).strip()
        if len(clean) > 30:
            puntos.append(clean)
        if len(puntos) >= n:
            break
    return puntos


_BROLL = {
    'Suplementación NutriHeal': (
        'SEC 0-3s: logo NutriHeal 360 animado fade-in sobre fondo negro. '
        'SEC 3-8s: close-up rotatorio del frasco de {producto} — 360° giro lento, iluminación de producto. '
        'SEC 8-15s: animación 3D del mecanismo de acción (molécula viajando al órgano objetivo). '
        'SEC 15-22s: infografía minimalista con los 3 beneficios clave del producto (texto aparece con slide-in). '
        'SEC 22-28s: sello INVIMA animado aparece en pantalla con partículas doradas. '
        'SEC 28-fin: CTA animado — ícono WhatsApp verde pulsante + número +57 314 708 8080.'
    ),
    'Medicina Alternativa': (
        'SEC 0-3s: logo NutriHeal 360 animado fade-in sobre fondo negro. '
        'SEC 3-10s: toma macro de viales homeopáticos y gránulos sobre superficie botánica — foco pull suave. '
        'SEC 10-18s: ilustración de iridodiagnosis animada — el iris se abre y el mapa de zonas aparece superpuesto. '
        'SEC 18-25s: manos del Dr. Mestizo examinando plantas medicinales en consultorio (close-up). '
        'SEC 25-fin: tarjeta de cierre — logo NutriHeal, dirección consultorio San Fernando Cali, CTA WhatsApp.'
    ),
    'Prevención Médica Diaria': (
        'SEC 0-3s: logo NutriHeal 360 animado fade-in. '
        'SEC 3-10s: flat-lay animado de vegetales y plantas medicinales con texto de estadística apareciendo (ej. "70% de las enfermedades crónicas son prevenibles"). '
        'SEC 10-18s: render 3D de cuerpo humano con órganos saludables iluminándose en verde esmeralda uno por uno. '
        'SEC 18-25s: familia colombiana saludable al aire libre — corte editorial, cámara lenta. '
        'SEC 25-fin: CTA animado — WhatsApp + "Agenda tu consulta hoy".'
    ),
    'Casos de Éxito / Trayectoria': (
        'SEC 0-3s: logo NutriHeal 360 animado fade-in. '
        'SEC 3-10s: tarjeta de testimonio con nombre de paciente (ficticio/genérico), foto de perfil anonimizada, y cita textual en scroll animado. '
        'SEC 10-18s: línea de tiempo animada de la trayectoria del Dr. Mestizo — diplomas y certificados aparecen con golden glow. '
        'SEC 18-25s: interior consultorio San Fernando — cámara lenta panorámica de pared de certificados. '
        'SEC 25-fin: CTA — WhatsApp + "Soy paciente, quiero agendar".'
    ),
}

_HEYGEN_PLAT = {
    'Facebook': {
        'encuadre':    'Plano americano clásico (cadera-cabeza), Dr. Mestizo ligeramente descentrado a la izquierda (regla de tercios), espacio derecho para texto overlay.',
        'fondo':       'Consultorio NutriHeal San Fernando Cali — escritorio médico de madera oscura en segundo plano, monstera grande a la izquierda, vitrina con frascos NutriHeal visibles y retroiluminados en verde esmeralda, diplomas enmarcados en dorado en la pared.',
        'iluminacion': 'Softbox 80×80cm frontal a 45° (luz principal, 5500K); reflector dorado lateral derecho (relleno cálido, 3200K 30%); backlight esmeralda (#1B7A4A) de bajo perfil detrás del sujeto para separarlo del fondo.',
        'voz':         'Tono: autoridad médica cálida y cercana. Ritmo: 130-140 ppm. Pausas dramatúrgicas después de afirmaciones clave (0.5s). Énfasis en sustantivos médicos importantes.',
        'formato':     '1:1 (Feed cuadrado) — 1080×1080 px. Captions en español con fuente sans-serif blanca, outline negro, tamaño 42pt, posición 10% desde el borde inferior.',
        'duracion':    '60-90 segundos. Los primeros 3s deben contener el hook sin intro.',
        'hook':        'Pregunta retórica directa al espectador relacionada con el tema (ej: "¿Sabías que el 80% de la fatiga crónica tiene solución natural?").',
        'cta':         'Últimos 8s: "Escríbeme al WhatsApp +57 314 708 8080 para una consulta personalizada." Overlay: botón WhatsApp animado verde.',
    },
    'Instagram': {
        'encuadre':    'Primer plano cerrado (hombros-cabeza), centrado en pantalla. El Dr. Mestizo habla directo a cámara. Encuadre seguro para zona 9:16 con notch superior e inferior.',
        'fondo':       'Consultorio minimalista — pared blanca con acento esmeralda, logo NutriHeal 360 visible en el fondo desenfocado (tamaño 15% del frame), planta en esquina izquierda.',
        'iluminacion': 'Luz natural simulada desde ventana lateral izquierda (difusa, 5600K); reflector blanco suave frontal para eliminar sombras bajo los ojos; sin backlight agresivo.',
        'voz':         'Tono: energético, cercano, conversacional. Ritmo: 150-160 ppm (más rápido para Reels). Primera oración sin pausa — engancha en los primeros 1.5s.',
        'formato':     '9:16 vertical (Reels) — 1080×1920 px. Captions en español, estilo TikTok: una palabra o frase corta a la vez, fuente bold blanca con sombra, posición central.',
        'duracion':    '30-45 segundos idealmente (máximo 60s). Sin intro, sin "hola soy el Dr. Mestizo" — empieza con el dato o la pregunta.',
        'hook':        'Primera oración: un dato sorprendente o una afirmación que genere curiosidad inmediata. Debe funcionar sin sonido (solo con captions).',
        'cta':         'Últimos 5s: "Enlace en mi bio para agendar" + ícono de flecha hacia arriba animado + sticker WhatsApp si la plataforma lo permite.',
    },
    'YouTube': {
        'encuadre':    'Plano medio-americano (cintura-cabeza), Dr. Mestizo centrado o ligeramente a la izquierda. Cámara a la altura de los ojos. Espacio de cabeza estándar (15% superior).',
        'fondo':       'Consultorio NutriHeal San Fernando Cali — estantería completa con libros médicos y tratados de nutrición, 3-4 plantas verdes variadas, pared con diploma principal enmarcado en dorado visible sobre el hombro derecho, iluminación cálida de lámpara de escritorio.',
        'iluminacion': 'Softbox 120×80cm a 45° frontal izquierdo (clave, 5500K, 100%); panel LED dorado (#C9A84C) lateral derecho como relleno cálido (40%); kicker esmeralda (#1B7A4A) desde abajo-atrás para separar del fondo (25%). Calidad YouTuber médico premium.',
        'voz':         'Tono: profesor universitario accesible. Ritmo: 120-130 ppm con pausas de 1s para puntos importantes. Entonación descendente al cierre de cada sección. Articulación clara.',
        'formato':     '16:9 horizontal — 1920×1080 px (Full HD). Captions opcionales (activar auto-subtítulos YouTube). Lower thirds animados para nombre del tema de cada sección.',
        'duracion':    '8-12 minutos. Estructura: Intro gancho (0-30s) → Contexto clínico (30s-2min) → Desarrollo en 3 puntos (2-9min) → Resumen + CTA (9-fin).',
        'hook':        'Los primeros 30 segundos: presentar el problema que el espectador reconoce, no el título del video. Ej: "Si llevas más de 6 meses con este síntoma, necesitas ver este video."',
        'cta':         'En minuto 7 (CTA de medio) y en los últimos 60s (CTA final): "Agenda tu consulta en el link de la descripción o escríbeme al WhatsApp +57 314 708 8080." Card final de YouTube visible.',
    },
}


def generar_heygen(copy, pilar, red, producto=None):
    """Construye la instrucción completa para HeyGen basada en el copy del post."""
    titulo = extraer_titulo_heygen(copy, pilar)
    puntos = extraer_puntos_guion(copy, n=3)
    meta   = _HEYGEN_PLAT.get(red, _HEYGEN_PLAT['YouTube'])

    broll_key = pilar if pilar in _BROLL else 'Medicina Alternativa'
    broll = _BROLL[broll_key]
    if producto and '{producto}' in broll:
        broll = broll.replace('{producto}', producto)
    elif '{producto}' in broll:
        broll = broll.replace('{producto} en close-up, ', '')

    guion_puntos = ' | '.join(f'({i+1}) {p}' for i, p in enumerate(puntos)) if puntos else titulo
    texto_pantalla = producto if ('Suplementación' in pilar and producto) else titulo[:70]

    partes = [
        '═══ INSTRUCCIÓN HEYGEN — DR. MESTIZO ═══',
        f'AVATAR: Dr. Mestizo — clonar con voz registrada del Dr. Mestizo (audio de referencia en carpeta /Audio_Referencia). Expresión neutral-cálida. Pestañeo natural activado.',
        f'TEMA: {titulo}.',
        f'PILAR: {pilar} | RED: {red}.',
        '',
        f'HOOK (apertura — primeros {3 if red == "Instagram" else 5}s): {meta["hook"]}',
        '',
        f'GUIÓN — PUNTOS CLAVE A DESARROLLAR:',
        guion_puntos,
        'Cierre con CTA específico (ver sección CTA abajo).',
        '',
        f'VOZ Y RITMO: {meta["voz"]}',
        '',
        f'ENCUADRE DE CÁMARA: {meta["encuadre"]}',
        f'FONDO / ESCENOGRAFÍA: {meta["fondo"]}',
        f'ILUMINACIÓN: {meta["iluminacion"]}',
        '',
        f'B-ROLL (inserts de video para cortar sobre el discurso):',
        broll,
        '',
        f'TEXTO EN PANTALLA PRINCIPAL: "{texto_pantalla}" — fuente sans-serif, color esmeralda (#1B7A4A), borde blanco, aparece en los primeros 5s del video.',
        f'LOWER THIRDS (si aplica): "Dr. Germán Mestizo | Médico Nutriólogo | Reg. Prof. XXXXXX" en el segundo 4-8.',
        '',
        f'CTA: {meta["cta"]}',
        f'FORMATO FINAL: {meta["formato"]}',
        f'DURACIÓN OBJETIVO: {meta["duracion"]}',
        '═══════════════════════════════════════',
    ]
    return ' '.join(partes)


def make_hyperlink_cell(ws, row, col, url, label):
    cell = ws.cell(row=row, column=col)
    cell.value = label
    cell.hyperlink = url
    cell.font = Font(color='0563C1', underline='single')
    cell.alignment = Alignment(vertical='top')

def actualizar_hoja_fuente(wb, nombre_hoja):
    ws = wb[nombre_hoja]

    # Asegurar encabezados en columnas 13 y 14
    ws.cell(row=1, column=13).value = '📦 Imagen Producto (Ref. 1 — ElevenLabs Image)'
    ws.cell(row=1, column=14).value = '🏷️ Logo NutriHeal (Ref. 2 — ElevenLabs Image)'
    for col in (13, 14):
        ws.cell(row=1, column=col).font = Font(bold=True)
        ws.cell(row=1, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions[get_column_letter(13)].width = 32
    ws.column_dimensions[get_column_letter(14)].width = 30

    # Red social de esta hoja (para HeyGen)
    red_nombre = nombre_hoja.replace('📘 ', '').replace('📷 ', '').replace('▶️ ', '').strip()

    filas_ok = 0
    filas_sin_map = 0
    for r in range(2, ws.max_row + 1):
        prompt_viejo = ws.cell(r, 9).value
        if not prompt_viejo:
            continue

        copy    = ws.cell(r, 8).value or ''
        pilar   = ws.cell(r, 4).value or ''

        # ── Col 9: prompt imagen ──────────────────────────────────────────────
        nuevo_p, nombre_prod = nuevo_prompt(str(prompt_viejo))
        ws.cell(r, 9).value = nuevo_p

        # ── Col 11: instrucción HeyGen dinámica ──────────────────────────────
        heygen_instr = generar_heygen(copy, pilar, red_nombre, producto=nombre_prod)
        ws.cell(r, 11).value = heygen_instr
        ws.cell(r, 11).alignment = Alignment(wrap_text=True, vertical='top')

        # ── Col 14: logo (siempre) ───────────────────────────────────────────
        make_hyperlink_cell(ws, r, 14, LOGO_URL, '🏷️ Logo NutriHeal')

        # ── Col 13: imagen del producto (solo Suplementación) ────────────────
        if nombre_prod:
            blob = blob_url_producto(nombre_prod)
            if blob:
                make_hyperlink_cell(ws, r, 13, blob, f'📦 {nombre_prod}')
            else:
                ws.cell(r, 13).value = f'⚠️ No encontrado en catálogo: {nombre_prod}'
                ws.cell(r, 13).alignment = Alignment(vertical='top')
        else:
            ws.cell(r, 13).value = 'N/A — imagen temática generativa'
            ws.cell(r, 13).alignment = Alignment(vertical='top')
            ws.cell(r, 13).font = Font(color='888888', italic=True)

        if nuevo_p == prompt_viejo and not PRODUCTO_RE_NEW.search(nuevo_p):
            filas_sin_map += 1
        filas_ok += 1

    print(f'  {nombre_hoja}: {filas_ok} filas actualizadas'
          + (f' ({filas_sin_map} sin mapeo)' if filas_sin_map else ''))

# ── Alt-text actualizado ─────────────────────────────────────────────────────
ALT_TEXT_MAP_NEW = {v: k2 for k2, v in {
    'Viales de vidrio y gránulos homeopáticos sobre fondo botánico, en tonos verde esmeralda y dorado — NutriHeal 360':
        'homeopathic glass vials and pellets on botanical background, blurred consultation room',
    'Primer plano del iris de un ojo humano con superposición de carta diagnóstica de iridología':
        'iris of a human eye close-up revealing health patterns with iridology diagnostic chart overlay',
    'Ilustración botánica médica de estilo vintage con marco dorado moderno':
        'vintage botanical medical illustration with elegant modern gold frame overlay',
    'Silueta humana con líneas de meridianos energéticos brillantes — visualización de sanación holística':
        'human silhouette with glowing energy meridian lines, holistic healing visualization',
    'Vegetales orgánicos frescos y plantas medicinales sobre pizarra oscura, estilo póster informativo de salud':
        'fresh organic vegetables and medicinal plants beautifully arranged on dark slate surface',
    'Ilustración anatómica del cuerpo humano con órganos saludables iluminados en verde esmeralda':
        'human body anatomical illustration with glowing healthy organs highlighted in emerald green',
    'Familia colombiana saludable al aire libre durante la hora dorada, energía de bienestar':
        'healthy Colombian family at golden hour outdoor setting, vibrant wellness energy',
    'Estructura de hélice de ADN con escudo protector superpuesto — concepto de prevención':
        'DNA helix structure with glowing protective shield overlay symbolizing prevention',
    'Paciente y médico colombianos en un momento de éxito dentro del consultorio, sonrisa cálida — pared de certificados, plantas verdes y diplomas enmarcados en dorado':
        'Colombian patient and doctor sharing a warm moment of success inside a consultation room',
}.items()}

ALT_TEXT_MAP_OLD = {
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, homeopathic glass vials and pellets on botanical background, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        'Viales de vidrio y gránulos homeopáticos sobre fondo botánico, en tonos verde esmeralda y dorado — NutriHeal 360',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, iris of human eye close-up revealing health patterns, iridology diagnostic chart overlay, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        'Primer plano del iris de un ojo humano con superposición de carta diagnóstica de iridología',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, vintage botanical medical illustration with modern gold frame overlay, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        'Ilustración botánica médica de estilo vintage con marco dorado moderno',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, human silhouette with glowing energy meridian lines, holistic healing visualization, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        'Silueta humana con líneas de meridianos energéticos brillantes — visualización de sanación holística',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, fresh organic vegetables and medicinal plants arranged on dark slate, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        'Vegetales orgánicos frescos y plantas medicinales sobre pizarra oscura, estilo póster informativo de salud',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, human body anatomical illustration with glowing healthy organs, emerald green highlights, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        'Ilustración anatómica del cuerpo humano con órganos saludables iluminados en verde esmeralda',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, healthy Colombian family at golden hour outdoor setting, vibrant wellness energy, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        'Familia colombiana saludable al aire libre durante la hora dorada, energía de bienestar',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, DNA helix structure with protective shield overlay, prevention concept, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        'Estructura de hélice de ADN con escudo protector superpuesto — concepto de prevención',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, Colombian patient and doctor moment of success in medical consultation room, warm smile, trust and gratitude emotion, San Fernando Cali medical office aesthetic, certificate wall background, emerald green plants, gold framed diplomas, before-after transformation concept':
        'Paciente y médico colombianos en un momento de éxito dentro del consultorio, sonrisa cálida — pared de certificados, plantas verdes y diplomas enmarcados en dorado',
}

def generar_alt_text(prompt):
    if not prompt:
        return ''
    m = PRODUCTO_RE_NEW.search(prompt)
    if m:
        producto = m.group(1)
        return (f'Foto de producto: frasco de {producto} sobre mármol oscuro con '
                f'brillo verde esmeralda, detalles dorados y hierbas medicinales — NutriHeal 360')
    m = PRODUCTO_RE_OLD.search(prompt)
    if m:
        producto = m.group(1)
        return (f'Foto de producto: frasco de {producto} sobre mármol oscuro con '
                f'brillo verde esmeralda, detalles dorados y hierbas medicinales — NutriHeal 360')
    return ALT_TEXT_MAP_OLD.get(prompt, ALT_TEXT_MAP_NEW.get(prompt, ''))

# ── Publer Import ────────────────────────────────────────────────────────────
PUBLER_HEADERS = [
    'Date - Intl. format or prompt', 'Text',
    'Link(s) - Separated by comma for FB carousels',
    'Media URL(s) - Separated by comma',
    'Title - For the video, pin, PDF ..',
    'Label(s) - Separated by comma',
    'Alt text(s) - Separated by ||',
    'Comment(s) - Separated by ||',
    'Pin board, FB album, or Google category',
    'Post subtype - I.e. story, reel, PDF ..',
    'CTA - For Facebook links or Google',
    'Reminder - For stories, reels, shorts, and TikToks',
]
EXTRA_HEADERS = [
    'Relación de Aspecto', 'Resolución Sugerida',
    'Duración Estimada', 'Pilar Temático', 'N° (ref. hoja origen)',
]
PLATFORM_META = {
    'Facebook':  {'sheet': '📘 Facebook',  'aspecto': '1:1 (Feed) / 16:9 (Video nativo)', 'resolucion': '1080x1080 / 1920x1080', 'duracion': '60-90 segundos (si es video nativo)', 'post_subtype': '', 'cta': 'Send Message', 'link': WHATSAPP_LINK},
    'Instagram': {'sheet': '📷 Instagram', 'aspecto': '9:16 (Vertical — Reels)',          'resolucion': '1080x1920',             'duracion': '30-60 segundos',                      'post_subtype': 'Reel',  'cta': '', 'link': ''},
    'YouTube':   {'sheet': '▶️ YouTube',   'aspecto': '16:9 (Horizontal)',                'resolucion': '1920x1080 (mínimo Full HD)', 'duracion': '8-12 minutos',                   'post_subtype': 'Video', 'cta': '', 'link': ''},
}
MEDIA_PENDIENTE = 'PENDIENTE — generar con ElevenLabs Image (imagen) / HeyGen (video avatar) y subir antes de programar'

from datetime import datetime
TITULO_RE = re.compile(r'"([^"]+)"')

def parse_fecha_hora(fecha_str, hora_str):
    fecha = datetime.strptime(str(fecha_str).strip(), '%d/%m/%Y')
    hora  = datetime.strptime(str(hora_str).strip(), '%I:%M %p')
    return fecha.replace(hour=hora.hour, minute=hora.minute).strftime('%Y-%m-%d %H:%M')

def construir_fila_publer(row, red):
    meta = PLATFORM_META[red]
    n, fecha, _dia, pilar, _red, _plat, hora, copy, prompt_mj = row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], row[8]
    hashtags = row[11] if len(row) > 11 else ''
    fecha_pub = parse_fecha_hora(fecha, hora)
    titulo = (TITULO_RE.search(copy or '').group(1) if red == 'YouTube' else '')
    alt = generar_alt_text(prompt_mj)
    publer = [
        fecha_pub, copy, meta['link'], MEDIA_PENDIENTE, titulo,
        pilar, alt, hashtags or '', '', meta['post_subtype'], meta['cta'], '',
    ]
    extra = [meta['aspecto'], meta['resolucion'], meta['duracion'], pilar, n]
    return publer + extra

def regenerar_publer_import(wb, red):
    meta = PLATFORM_META[red]
    src = wb[meta['sheet']]
    rows = list(src.iter_rows(min_row=2, max_col=12, values_only=True))

    nombre_hoja = f'📤 Publer Import - {red}'
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    ws = wb.create_sheet(nombre_hoja)

    headers = PUBLER_HEADERS + EXTRA_HEADERS
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.freeze_panes = 'A2'

    for row in rows:
        ws.append(construir_fila_publer(row, red))

    anchos = [16, 50, 22, 50, 40, 22, 45, 30, 18, 16, 16, 18, 26, 22, 24, 22, 10]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho
    for r in range(2, ws.max_row + 1):
        for c in (2, 7, 8):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical='top')

    return ws, len(rows)

def exportar_csv(ws, red, total_filas):
    out = os.path.join(BASE_DIR, f'publer_import_{red.lower()}.csv')
    with open(out, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(PUBLER_HEADERS)
        for r in range(2, total_filas + 2):
            fila = [ws.cell(row=r, column=c).value or '' for c in range(1, len(PUBLER_HEADERS) + 1)]
            writer.writerow(fila)
    return out

def main():
    wb = openpyxl.load_workbook(SRC_XLSX)

    print('=== Actualizando hojas fuente ===')
    for hoja in ['📘 Facebook', '📷 Instagram', '▶️ YouTube']:
        actualizar_hoja_fuente(wb, hoja)

    print('\n=== Regenerando hojas Publer Import ===')
    for red in ['Facebook', 'Instagram', 'YouTube']:
        ws, total = regenerar_publer_import(wb, red)
        csv_path = exportar_csv(ws, red, total)
        print(f'  {red}: {total} filas → hoja "📤 Publer Import - {red}" + {os.path.basename(csv_path)}')

    wb.save(SRC_XLSX)
    print(f'\n✅ Guardado: {os.path.basename(SRC_XLSX)}')
    print(f'📁 Carpeta Marketing: {os.path.join(BASE_DIR, "Marketing")}')

if __name__ == '__main__':
    main()
