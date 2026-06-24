# -*- coding: utf-8 -*-
"""
Genera hojas "Publer Import" (alineadas al CSV Template for Publer) dentro del
archivo de auditoria NutriHeal360_DrMestizo_Auditoria_Contenido_3Meses_2026.xlsx,
y exporta los 3 CSV listos para subir a Publer.

Agrega ademas columnas de produccion que el template de Publer no contempla
pero que el equipo necesita: Relacion de Aspecto, Resolucion Sugerida,
Duracion Estimada, Pilar Tematico y N° de referencia.
"""
import csv
import os
import re
import sys
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SRC_XLSX = os.path.join(BASE_DIR, 'NutriHeal360_DrMestizo_Auditoria_Contenido_3Meses_2026.xlsx')
WHATSAPP_LINK = 'https://wa.me/573147088080'

PUBLER_HEADERS = [
    'Date - Intl. format or prompt',
    'Text',
    'Link(s) - Separated by comma for FB carousels',
    'Media URL(s) - Separated by comma',
    'Title - For the video, pin, PDF ..',
    'Label(s) - Separated by comma',
    'Alt text(s) - Separated by ||',
    'Comment(s) - Separated by ||',
    'Pin board, FB album, or Google category',
    'Post subtype - I.e. story, reel, PDF ..',
    'CTA - For Facebook links or Google',
    'Reminder - For stories, reels, shorts, and TikToks',
]

EXTRA_HEADERS = [
    'Relación de Aspecto',
    'Resolución Sugerida',
    'Duración Estimada',
    'Pilar Temático',
    'N° (ref. hoja origen)',
]

PLATFORM_META = {
    'Facebook': {
        'sheet': '📘 Facebook',
        'aspecto': '1:1 (Feed) / 16:9 (Video nativo)',
        'resolucion': '1080x1080 / 1920x1080',
        'duracion': '60-90 segundos (si es video nativo)',
        'post_subtype': '',
        'cta': 'Send Message',
        'link': WHATSAPP_LINK,
    },
    'Instagram': {
        'sheet': '📷 Instagram',
        'aspecto': '9:16 (Vertical — Reels)',
        'resolucion': '1080x1920',
        'duracion': '30-60 segundos',
        'post_subtype': 'Reel',
        'cta': '',
        'link': '',
    },
    'YouTube': {
        'sheet': '▶️ YouTube',
        'aspecto': '16:9 (Horizontal)',
        'resolucion': '1920x1080 (mínimo Full HD)',
        'duracion': '8-12 minutos',
        'post_subtype': 'Video',
        'cta': '',
        'link': '',
    },
}

MEDIA_PENDIENTE = 'PENDIENTE — generar con Midjourney/HeyGen y subir antes de programar'

# Traducciones manuales de los 9 prompts no-templados (Medicina Alternativa,
# Prevencion Medica Diaria y Casos de Exito). El pilar "Suplementacion
# NutriHeal" se genera con plantilla (ver generar_alt_text).
ALT_TEXT_MAP = {
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, homeopathic glass vials and pellets on botanical background, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        'Viales de vidrio y gránulos homeopáticos sobre fondo botánico, en tonos verde esmeralda y dorado — NutriHeal 360',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, iris of human eye close-up revealing health patterns, iridology diagnostic chart overlay, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        'Primer plano del iris de un ojo humano con superposición de carta diagnóstica de iridología',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, vintage botanical medical illustration with modern gold frame overlay, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        'Ilustración botánica médica de estilo vintage con marco dorado moderno',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, human silhouette with glowing energy meridian lines, holistic healing visualization, Colombian doctor consultation room background blurred, emerald and gold tones, trust and science emotion':
        'Silueta humana con líneas de meridianos energéticos brillantes — visualización de sanación holística',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, fresh organic vegetables and medicinal plants arranged on dark slate, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        'Vegetales orgánicos frescos y plantas medicinales sobre pizarra oscura, estilo póster informativo de salud',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, human body anatomical illustration with glowing healthy organs, emerald green highlights, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        'Ilustración anatómica del cuerpo humano con órganos saludables iluminados en verde esmeralda',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, healthy Colombian family at golden hour outdoor setting, vibrant wellness energy, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        'Familia colombiana saludable al aire libre durante la hora dorada, energía de bienestar',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, DNA helix structure with protective shield overlay, prevention concept, informative health poster style, Colombian demographic, emerald green and gold accents, hope and vitality emotion':
        'Estructura de hélice de ADN con escudo protector superpuesto — concepto de prevención',
    'Photorealistic professional medical marketing image, emerald green (#1B7A4A) and premium gold (#C9A84C) color palette, luxury health brand aesthetic, ultra-detailed, 8K resolution, soft cinematic lighting, Colombian patient and doctor moment of success in medical consultation room, warm smile, trust and gratitude emotion, San Fernando Cali medical office aesthetic, certificate wall background, emerald green plants, gold framed diplomas, before-after transformation concept':
        'Paciente y médico colombianos en un momento de éxito dentro del consultorio, sonrisa cálida — pared de certificados, plantas verdes y diplomas enmarcados en dorado',
}

PRODUCTO_RE = re.compile(r'of \"([^\"]+)\" on a dark marble')


def generar_alt_text(prompt):
    if not prompt:
        return ''
    m = PRODUCTO_RE.search(prompt)
    if m:
        producto = m.group(1)
        return (f'Foto de producto: frasco de {producto} sobre mármol oscuro con '
                f'brillo verde esmeralda, detalles dorados y hierbas medicinales — NutriHeal 360')
    return ALT_TEXT_MAP.get(prompt, '')


def parse_fecha_hora(fecha_str, hora_str):
    """'22/06/2026' + '7:00 AM' -> '2026-06-22 07:00'"""
    fecha = datetime.strptime(str(fecha_str).strip(), '%d/%m/%Y')
    hora = datetime.strptime(str(hora_str).strip(), '%I:%M %p')
    combinado = fecha.replace(hour=hora.hour, minute=hora.minute)
    return combinado.strftime('%Y-%m-%d %H:%M')


TITULO_RE = re.compile(r'"([^"]+)"')


def extraer_titulo(copy_texto):
    m = TITULO_RE.search(copy_texto or '')
    return m.group(1) if m else ''


def construir_fila(row, red):
    meta = PLATFORM_META[red]
    n, fecha, _dia, pilar, _red, _plat, hora, copy, prompt_mj, _eleven, _heygen, hashtags = row

    fecha_publer = parse_fecha_hora(fecha, hora)
    titulo = extraer_titulo(copy) if red == 'YouTube' else ''
    alt_text = generar_alt_text(prompt_mj)

    publer_cols = [
        fecha_publer,                  # Date
        copy,                          # Text
        meta['link'],                  # Link(s)
        MEDIA_PENDIENTE,                # Media URL(s)
        titulo,                        # Title
        pilar,                         # Label(s)
        alt_text,                      # Alt text(s)
        hashtags or '',                # Comment(s)
        '',                            # Pin board...
        meta['post_subtype'],          # Post subtype
        meta['cta'],                   # CTA
        '',                            # Reminder
    ]
    extra_cols = [
        meta['aspecto'],
        meta['resolucion'],
        meta['duracion'],
        pilar,
        n,
    ]
    return publer_cols + extra_cols


def escribir_hoja(wb, red):
    meta = PLATFORM_META[red]
    src = wb[meta['sheet']]
    rows = list(src.iter_rows(min_row=2, values_only=True))

    nombre_hoja = f'📤 Publer Import - {red}'
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    ws = wb.create_sheet(nombre_hoja)

    headers = PUBLER_HEADERS + EXTRA_HEADERS
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws.freeze_panes = 'A2'

    for row in rows:
        ws.append(construir_fila(row, red))

    anchos = [16, 50, 22, 32, 40, 22, 45, 30, 18, 16, 16, 18, 26, 22, 24, 22, 10]
    for i, ancho in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    for r in range(2, ws.max_row + 1):
        for c in (2, 7, 8):  # Text, Alt text, Comment(s)
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical='top')

    return ws, len(rows)


def exportar_csv(ws, red, total_filas):
    """Exporta SOLO las 12 columnas Publer (sin las columnas extra de produccion)."""
    out_path = os.path.join(BASE_DIR, f'publer_import_{red.lower()}.csv')
    with open(out_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(PUBLER_HEADERS)
        for r in range(2, total_filas + 2):
            fila = [ws.cell(row=r, column=c).value or '' for c in range(1, len(PUBLER_HEADERS) + 1)]
            writer.writerow(fila)
    return out_path


def actualizar_guia(wb):
    ws = wb['📖 Guía de Uso']
    fila = ws.max_row + 2
    notas = [
        ('HOJAS "PUBLER IMPORT" (NUEVO)', '', '', '', ''),
        ('Hoja', 'Descripción', 'Posts', 'Observaciones', ''),
        ('📤 Publer Import - Facebook', 'Estructura exacta del CSV Template for Publer + columnas de producción', '92',
         'Link(s)=WhatsApp, CTA=Send Message', ''),
        ('📤 Publer Import - Instagram', 'Igual estructura, Post subtype=Reel', '92',
         'Hashtags van en Comment(s), no en Text', ''),
        ('📤 Publer Import - YouTube', 'Igual estructura, Post subtype=Video, Title extraído del copy', '92',
         'Revisar hashtag #Shorts heredado (el video es de 8-12 min, no Short)', ''),
        ('', '', '', '', ''),
        ('GAP PENDIENTE', '', '', '', ''),
        ('Media URL(s) queda en "PENDIENTE" en las 276 filas.', 'Generar las imágenes/videos con Midjourney y HeyGen '
         'usando los prompts de las hojas originales, subir los archivos, y pegar la URL final en esa columna '
         'ANTES de exportar a Publer.', '', '', ''),
        ('', '', '', '', ''),
        ('CÓMO EXPORTAR A PUBLER', '', '', '', ''),
        ('1. Completar Media URL(s) en la hoja correspondiente.', '', '', '', ''),
        ('2. Re-ejecutar generar_publer_import.py para regenerar los 3 CSV con los datos finales.', '', '', '', ''),
        ('3. En Publer: Bulk Import > seleccionar cuenta > subir publer_import_[red].csv', '', '', '', ''),
    ]
    for fila_datos in notas:
        ws.append(fila_datos)


def main():
    if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')

    wb = openpyxl.load_workbook(SRC_XLSX)

    resumen = []
    for red in ['Facebook', 'Instagram', 'YouTube']:
        ws, total = escribir_hoja(wb, red)
        csv_path = exportar_csv(ws, red, total)
        resumen.append((red, total, csv_path))

    actualizar_guia(wb)
    wb.save(SRC_XLSX)

    print('=== Listo ===')
    for red, total, csv_path in resumen:
        print(f'{red}: {total} filas -> hoja "📤 Publer Import - {red}" + {os.path.basename(csv_path)}')


if __name__ == '__main__':
    main()
