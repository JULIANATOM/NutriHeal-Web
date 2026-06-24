# -*- coding: utf-8 -*-
"""
Completa el texto truncado en los copies de Instagram y YouTube.
Hay exactamente 4 fragmentos truncados distintos, todos en Prevención Médica Diaria.
El texto completo se extrae del copy de Facebook (fuente de verdad).
"""
import sys, openpyxl
from openpyxl.styles import Alignment
sys.stdout.reconfigure(encoding='utf-8')

SRC = 'NutriHeal360_DrMestizo_Auditoria_Contenido_3Meses_2026.xlsx'

# ── Reemplazos exactos: (fragmento truncado) -> (versión completa) ───────────
# Extraídos de los copies de Facebook para los mismos pilar/fecha
FIXES = {
    # Instagram trunca después de "Y lo más "
    'Y lo más ..."':
        'Y lo más importante: es completamente prevenible con alimentación y suplementación correcta."',

    # Instagram trunca después de "Genera inf"
    'Genera inf..."':
        'Genera inflamación, destruye la microbiota y agota el páncreas lentamente."',

    # YouTube trunca después de "es compl"
    'es compl..."':
        'es completamente prevenible con alimentación y suplementación correcta."',

    # YouTube trunca después de "destruye l"
    'destruye l..."':
        'destruye la microbiota y agota el páncreas lentamente."',
}

def corregir_copy(text):
    for viejo, nuevo in FIXES.items():
        if viejo in text:
            text = text.replace(viejo, nuevo)
    return text

def main():
    wb = openpyxl.load_workbook(SRC)
    total = 0

    for hoja in ['📘 Facebook', '📷 Instagram', '▶️ YouTube']:
        ws = wb[hoja]
        corregidas = 0
        for r in range(2, ws.max_row + 1):
            copy_orig = ws.cell(r, 8).value
            if not copy_orig:
                continue
            copy_nuevo = corregir_copy(str(copy_orig))
            if copy_nuevo != str(copy_orig):
                ws.cell(r, 8).value = copy_nuevo
                ws.cell(r, 8).alignment = Alignment(wrap_text=True, vertical='top')
                corregidas += 1
        print(f'  {hoja}: {corregidas} copys completados')
        total += corregidas

    wb.save(SRC)
    print(f'\n✅ Total corregidos: {total} | Guardado: {SRC}')

if __name__ == '__main__':
    main()
